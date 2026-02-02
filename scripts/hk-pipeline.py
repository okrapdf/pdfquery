#!/usr/bin/env python3
"""
HK Financial Report Extraction Pipeline

Downloads a PDF from URL, extracts financial tables using Qwen VL via OpenRouter,
and converts to Excel.

Usage:
  # Full pipeline (scout + extract + xlsx)
  python scripts/hk-pipeline.py --url "https://example.com/report.pdf" --company "Galaxy Entertainment" --stock-code 00640 --year 2024

  # Extract specific pages only (skip scout)
  python scripts/hk-pipeline.py --url "https://example.com/report.pdf" --pages 45,46,47,48,49,50 --company "Galaxy Entertainment"

  # From local PDF
  python scripts/hk-pipeline.py --pdf /path/to/report.pdf --company "Galaxy Entertainment" --stock-code 00640 --year 2024

Env:
  OPENROUTER_API_KEY - Required
"""

import os
import sys
import json
import base64
import time
import re
import argparse
import tempfile
from pathlib import Path
from datetime import datetime
from io import BytesIO

import fitz  # PyMuPDF
import httpx
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY")
MODEL = "qwen/qwen3-vl-235b-a22b-instruct"
INPUT_COST_PER_M = 0.20
OUTPUT_COST_PER_M = 1.20

# HK Annual Report Template - what tables to look for
HK_TEMPLATE = {
    "income_statement": {
        "search_terms": ["profit or loss", "income statement", "revenue", "損益表", "consolidated statement of profit"],
        "prompt": (
            "Extract the FULL consolidated income statement / profit or loss table from this page. "
            "Include ALL line items, subtotals, comparative columns (current year and prior year), "
            "and the final net profit/loss figure. Return as a clean markdown table. "
            "Preserve all numbers exactly as shown. If amounts are in millions, note that in the header."
        ),
        "sheet_name": "Income Statement",
        "required": True,
    },
    "balance_sheet": {
        "search_terms": ["financial position", "balance sheet", "財務狀況表", "consolidated statement of financial position"],
        "prompt": (
            "Extract the FULL consolidated balance sheet / statement of financial position table. "
            "Include current/non-current assets, current/non-current liabilities, total equity, "
            "and all line items with comparative columns. Return as a clean markdown table."
        ),
        "sheet_name": "Balance Sheet",
        "required": True,
    },
    "cash_flow": {
        "search_terms": ["cash flows", "cash flow statement", "現金流量表", "consolidated statement of cash flows"],
        "prompt": (
            "Extract the FULL consolidated cash flow statement. "
            "Include operating, investing, and financing activities with all line items "
            "and comparative columns. Return as a clean markdown table."
        ),
        "sheet_name": "Cash Flow",
        "required": True,
    },
    "segment_revenue": {
        "search_terms": ["segment", "revenue by", "business review", "revenue analysis"],
        "prompt": (
            "Extract the segment revenue / business segment breakdown table. "
            "Include all segments and their revenue figures. Return as a clean markdown table."
        ),
        "sheet_name": "Segment Revenue",
        "required": False,
    },
    "five_year_summary": {
        "search_terms": ["financial summary", "five year", "highlights", "財務摘要", "financial highlights"],
        "prompt": (
            "Extract the multi-year financial summary/highlights table. "
            "Include all years and metrics shown. Return as a clean markdown table."
        ),
        "sheet_name": "Financial Summary",
        "required": False,
    },
}

# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def download_pdf(url: str) -> Path:
    """Download PDF from URL to temp file."""
    print(f"Downloading PDF from {url[:80]}...")
    with httpx.Client(follow_redirects=True, timeout=60.0) as client:
        resp = client.get(url)
        resp.raise_for_status()

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    tmp.write(resp.content)
    tmp.close()
    print(f"  → {len(resp.content) / 1024 / 1024:.1f} MB saved to {tmp.name}")
    return Path(tmp.name)


def page_to_base64(page: fitz.Page, dpi: int = 150) -> str:
    """Convert PDF page to base64 PNG."""
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat)
    img_bytes = pix.tobytes("png")
    return base64.b64encode(img_bytes).decode("utf-8")


def get_page_text(page: fitz.Page) -> str:
    """Get raw text from a PDF page."""
    return page.get_text("text")


# ---------------------------------------------------------------------------
# OpenRouter / Qwen VL
# ---------------------------------------------------------------------------

def call_qwen_vl(client: httpx.Client, image_b64: str, prompt: str) -> dict:
    """Send image to Qwen VL via OpenRouter and return response."""
    payload = {
        "model": MODEL,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}},
                ],
            },
        ],
        "temperature": 0.1,
        "max_tokens": 8000,
        "provider": {"zdr": True, "data_collection": "deny", "sort": "throughput"},
    }

    resp = client.post(
        "https://openrouter.ai/api/v1/chat/completions",
        json=payload,
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
        },
        timeout=120.0,
    )
    resp.raise_for_status()
    data = resp.json()

    content = data["choices"][0]["message"]["content"].strip()
    usage = data.get("usage", {})

    return {
        "content": content,
        "input_tokens": usage.get("prompt_tokens", 0),
        "output_tokens": usage.get("completion_tokens", 0),
    }


# ---------------------------------------------------------------------------
# Table Scout — find which pages contain target tables
# ---------------------------------------------------------------------------

def _find_auditor_report_page(doc: fitz.Document) -> int | None:
    """
    Find the page of the Independent Auditor's Report.
    This marks the start of the audited financial statements section.
    Returns 0-indexed page number or None.
    """
    auditor_patterns = [
        "independent auditor",
        "report of the auditors",
        "auditor's report",
        "auditors' report",
        "independent auditors",
        "獨立核數師報告",
        "獨立審計師報告",
    ]
    for page_num in range(len(doc)):
        text = get_page_text(doc[page_num]).lower()
        for pat in auditor_patterns:
            if pat.lower() in text:
                return page_num
    return None


def _is_header_match(page_text: str, table_type: str) -> bool:
    """
    Check if the page header (first ~500 chars) contains a consolidated statement header
    matching the table type. This avoids false positives from TOC/overview pages where
    keywords appear in body text.
    """
    # Header patterns: consolidated statement titles that appear at the TOP of the page.
    # These are the actual financial statement page titles in HKEX annual reports.
    header_patterns = {
        "income_statement": [
            "consolidated statement of profit or loss",
            "consolidated income statement",
            "consolidated statement of profit",
            "consolidated statement of comprehensive income",
            "合併損益表",
        ],
        "balance_sheet": [
            "consolidated statement of financial position",
            "consolidated balance sheet",
            "合併財務狀況表",
        ],
        "cash_flow": [
            "consolidated statement of cash flows",
            "consolidated cash flow statement",
            "合併現金流量表",
        ],
        "segment_revenue": [
            "segment information",
            "segment revenue",
            "revenue by segment",
            "business segment",
        ],
        "five_year_summary": [
            "financial summary",
            "five year financial summary",
            "five-year financial summary",
            "financial highlights",
            "ten year financial summary",
            "ten-year financial summary",
            "財務摘要",
        ],
    }
    patterns = header_patterns.get(table_type, [])
    header = page_text[:500].lower()
    return any(p.lower() in header for p in patterns)


def scout_tables(doc: fitz.Document, client: httpx.Client) -> dict:
    """
    Scout the PDF to find pages containing target financial tables.

    Strategy:
    1. Find the "Independent Auditor's Report" page — financial statements follow after it.
    2. For core statements (income, balance, cash flow), only search AFTER the auditor page
       and match on page headers (first ~500 chars) to avoid TOC/overview false positives.
    3. For optional tables (segment, summary), search the full document with header matching.

    Returns: {table_type: [page_numbers]}
    """
    total_pages = len(doc)
    print(f"\n--- SCOUTING {total_pages} pages ---")

    # Step 1: Find Auditor's Report — anchor for financial statements section
    auditor_page = _find_auditor_report_page(doc)
    if auditor_page is not None:
        # Financial statements typically start 2-5 pages after auditor's report
        fs_start = auditor_page + 1
        print(f"  Auditor's Report found at page {auditor_page + 1}")
        print(f"  Searching for financial statements from page {fs_start + 1}")
    else:
        # Fallback: search the back half of the document (financial statements are always in the back)
        fs_start = total_pages // 2
        print(f"  Auditor's Report NOT found — searching from page {fs_start + 1} (back half)")

    core_types = {"income_statement", "balance_sheet", "cash_flow"}

    # Step 2: Header-based search
    found = {}
    for table_type, config in HK_TEMPLATE.items():
        found[table_type] = []
        # Core financial statements: only look after auditor page
        search_start = fs_start if table_type in core_types else 0
        search_end = total_pages

        for page_idx in range(search_start, search_end):
            text = get_page_text(doc[page_idx])
            if _is_header_match(text, table_type):
                found[table_type].append(page_idx + 1)  # 1-indexed

    # Step 3: For balance sheet, include the NEXT page if found (often spans 2 pages)
    if found.get("balance_sheet"):
        first_bs_page = found["balance_sheet"][0]
        if first_bs_page + 1 <= total_pages and (first_bs_page + 1) not in found["balance_sheet"]:
            found["balance_sheet"].append(first_bs_page + 1)
            found["balance_sheet"].sort()

    # Print results
    for table_type, pages in found.items():
        label = HK_TEMPLATE[table_type]["sheet_name"]
        if pages:
            print(f"  {label}: pages {pages[:5]}{'...' if len(pages) > 5 else ''}")
        else:
            print(f"  {label}: NOT FOUND")

    # Step 4: Return results — cap at 3 pages per type
    result = {}
    for table_type, pages in found.items():
        result[table_type] = pages[:3] if pages else []

    return result


# ---------------------------------------------------------------------------
# Extract tables from specific pages
# ---------------------------------------------------------------------------

def extract_table(doc: fitz.Document, client: httpx.Client, page_num: int, table_type: str) -> dict:
    """Extract a specific table from a page using VLM."""
    config = HK_TEMPLATE[table_type]
    page = doc[page_num - 1]  # 0-indexed

    image_b64 = page_to_base64(page, dpi=200)  # Higher DPI for financial tables
    image_size_kb = len(image_b64) * 3 / 4 / 1024

    print(f"  Page {page_num} ({table_type}) [{image_size_kb:.0f}KB]...", end=" ", flush=True)

    result = call_qwen_vl(client, image_b64, config["prompt"])

    has_table = "NO_TABLES" not in result["content"].upper() and "|" in result["content"]
    status = "✓" if has_table else "✗"
    print(f"{status} | in:{result['input_tokens']} out:{result['output_tokens']}")

    return {
        "table_type": table_type,
        "page_number": page_num,
        "markdown": result["content"] if has_table else "",
        "has_table": has_table,
        "input_tokens": result["input_tokens"],
        "output_tokens": result["output_tokens"],
    }


# ---------------------------------------------------------------------------
# Markdown → Excel
# ---------------------------------------------------------------------------

def parse_markdown_table(md: str) -> list[list[str]]:
    """Parse a markdown table into rows of cells."""
    lines = md.strip().split("\n")
    rows = []
    for line in lines:
        line = line.strip()
        if not line.startswith("|"):
            continue
        # Skip separator lines (|---|---|)
        if re.match(r"^\|[\s\-:]+\|", line):
            continue
        cells = [c.strip() for c in line.split("|")]
        # Remove empty first/last from leading/trailing pipes
        if cells and cells[0] == "":
            cells = cells[1:]
        if cells and cells[-1] == "":
            cells = cells[:-1]
        if cells:
            rows.append(cells)
    return rows


def create_excel(tables: list[dict], company_name: str, report_year: str, output_path: Path):
    """Convert extracted markdown tables to a formatted Excel workbook."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for table in tables:
        if not table.get("markdown") or not table.get("has_table"):
            continue

        table_type = table["table_type"]
        sheet_name = HK_TEMPLATE[table_type]["sheet_name"]
        rows = parse_markdown_table(table["markdown"])

        if not rows:
            continue

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        # Title row
        ws.cell(row=1, column=1, value=f"{company_name} - {sheet_name} ({report_year})")
        ws.cell(row=1, column=1).font = header_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(rows[0]), 1))

        # Source page reference
        ws.cell(row=2, column=1, value=f"Source: Page {table['page_number']}")
        ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="808080")

        # Data rows (starting at row 4)
        start_row = 4
        for i, row in enumerate(rows):
            for j, cell_value in enumerate(row):
                cell = ws.cell(row=start_row + i, column=j + 1, value=cell_value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                # Header row styling
                if i == 0:
                    cell.fill = header_fill
                    cell.font = header_font_white

        # Auto-fit column widths (approximate)
        for col_idx in range(1, len(rows[0]) + 1 if rows else 1):
            max_len = 0
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value="No tables were extracted.")

    wb.save(output_path)
    print(f"\n  Excel saved: {output_path} ({len(wb.sheetnames)} sheets)")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="HK Financial Report Extraction Pipeline")
    parser.add_argument("--url", help="PDF URL to download")
    parser.add_argument("--pdf", help="Local PDF path")
    parser.add_argument("--company", required=True, help="Company name (e.g. 'Galaxy Entertainment')")
    parser.add_argument("--stock-code", default="", help="Stock code (e.g. '00640')")
    parser.add_argument("--year", default="2024", help="Report year")
    parser.add_argument("--pages", help="Comma-separated page numbers to extract (skip scouting)")
    parser.add_argument("--page-map", help='JSON mapping table_type to pages, e.g. \'{"income_statement":[72],"balance_sheet":[74,75]}\'')

    parser.add_argument("--output-dir", help="Output directory (default: scripts/output/hk_<timestamp>)")
    parser.add_argument("--dpi", type=int, default=200, help="Image DPI for extraction")
    args = parser.parse_args()

    if not OPENROUTER_API_KEY:
        print("Error: OPENROUTER_API_KEY not set. Run: export OPENROUTER_API_KEY=...")
        sys.exit(1)

    if not args.url and not args.pdf:
        print("Error: Provide --url or --pdf")
        sys.exit(1)

    # Setup output dir
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir) if args.output_dir else Path(f"scripts/output/hk_{args.stock_code}_{timestamp}")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Get PDF
    if args.url:
        pdf_path = download_pdf(args.url)
        cleanup_pdf = True
    else:
        pdf_path = Path(args.pdf)
        cleanup_pdf = False

    if not pdf_path.exists():
        print(f"Error: PDF not found: {pdf_path}")
        sys.exit(1)

    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    print(f"\nPDF: {pdf_path.name} ({total_pages} pages)")
    print(f"Company: {args.company} ({args.stock_code})")
    print(f"Year: {args.year}")
    print(f"Output: {output_dir}")
    print("=" * 60)

    client = httpx.Client()
    total_input_tokens = 0
    total_output_tokens = 0
    all_tables = []
    start_time = time.time()

    try:
        # Phase 1: Scout or use provided pages
        if args.page_map:
            page_map = json.loads(args.page_map)
            print(f"\nUsing explicit page map:")
            for tt, pages in page_map.items():
                print(f"  {tt}: {pages}")
        elif args.pages:
            page_nums = [int(p.strip()) for p in args.pages.split(",")]
            page_map = {}
            table_types = list(HK_TEMPLATE.keys())
            for i, p in enumerate(page_nums):
                tt = table_types[i % len(table_types)] if i < len(table_types) else f"table_{i}"
                if tt not in page_map:
                    page_map[tt] = []
                page_map[tt].append(p)
            print(f"\nUsing provided pages: {page_nums}")
        else:
            # Scout for tables
            page_map = scout_tables(doc, client)

        # Phase 2: Extract tables
        print(f"\n--- EXTRACTING TABLES ---")
        for table_type, pages in page_map.items():
            if not pages:
                print(f"  {table_type}: no pages found, skipping")
                continue
            for page_num in pages:
                if page_num > total_pages:
                    print(f"  Page {page_num} exceeds total ({total_pages}), skipping")
                    continue
                result = extract_table(doc, client, page_num, table_type)
                all_tables.append(result)
                total_input_tokens += result["input_tokens"]
                total_output_tokens += result["output_tokens"]
                time.sleep(0.5)  # Rate limit

        # Phase 3: Save markdown outputs
        print(f"\n--- SAVING RESULTS ---")
        for table in all_tables:
            if table["has_table"]:
                md_file = output_dir / f"{table['table_type']}_p{table['page_number']}.md"
                with open(md_file, "w", encoding="utf-8") as f:
                    f.write(f"# {HK_TEMPLATE.get(table['table_type'], {}).get('sheet_name', table['table_type'])}\n")
                    f.write(f"Page {table['page_number']}\n\n")
                    f.write(table["markdown"])
                print(f"  Saved: {md_file.name}")

        # Phase 4: Create Excel
        xlsx_path = output_dir / f"{args.stock_code}_{args.company.replace(' ', '_')}_{args.year}.xlsx"
        create_excel(all_tables, args.company, args.year, xlsx_path)

        # Phase 5: Save metadata
        elapsed = time.time() - start_time
        input_cost = (total_input_tokens / 1_000_000) * INPUT_COST_PER_M
        output_cost = (total_output_tokens / 1_000_000) * OUTPUT_COST_PER_M
        total_cost = input_cost + output_cost

        meta = {
            "company_name": args.company,
            "stock_code": args.stock_code,
            "report_year": args.year,
            "pdf_url": args.url or str(args.pdf),
            "total_pages": total_pages,
            "tables_extracted": sum(1 for t in all_tables if t["has_table"]),
            "tables_attempted": len(all_tables),
            "elapsed_seconds": round(elapsed, 1),
            "input_tokens": total_input_tokens,
            "output_tokens": total_output_tokens,
            "total_cost_usd": round(total_cost, 4),
            "xlsx_path": str(xlsx_path),
            "model": MODEL,
            "timestamp": timestamp,
            "tables": [
                {
                    "type": t["table_type"],
                    "page": t["page_number"],
                    "has_table": t["has_table"],
                    "input_tokens": t["input_tokens"],
                    "output_tokens": t["output_tokens"],
                }
                for t in all_tables
            ],
        }

        meta_file = output_dir / "meta.json"
        with open(meta_file, "w") as f:
            json.dump(meta, f, indent=2)

        # Summary
        print(f"""
{'=' * 60}
EXTRACTION COMPLETE
{'=' * 60}
Company: {args.company} ({args.stock_code})
Year: {args.year}
Pages processed: {len(all_tables)}
Tables extracted: {meta['tables_extracted']}/{meta['tables_attempted']}
Time: {elapsed:.1f}s

Tokens: {total_input_tokens:,} in + {total_output_tokens:,} out
Cost: ${total_cost:.4f}

Output: {output_dir}
Excel: {xlsx_path}
{'=' * 60}""")

    finally:
        client.close()
        doc.close()
        if cleanup_pdf:
            os.unlink(pdf_path)


if __name__ == "__main__":
    main()
